import base64
import datetime
import json
import os

import xlrd
from apps.good_apply.models import Organization
from apps.good_apply.serializers import PreApplySerializers
from apps.tools.generate_can_not_apply_excel import \
    generate_can_not_apply_excel
from apps.tools.param_check import get_error_message
from apps.tools.response import get_result
from apps.tools.tool_get_import_file import tool_get_import_file
from apps.tools.tool_get_xlsx_excel_data import tool_get_xlsx_excel_data
from apps.tools.tool_valid_user_in_the_org import valid_user_in_the_org
from apps.utils.enums import StatusEnums
from apps.utils.exceptions import BusinessException
from bkstorages.backends.bkrepo import BKRepoStorage
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
from xlrd import xldate_as_tuple


@require_POST
def analysis_apply_excel(request):
    """解析excel文件数据"""
    def handle_excel_data(rows, CANNOT_APPLY):
        """处理传入的列表数据，判断是否加入部门所需物资表"""
        validated_list = []  # 存放GroupApply对象
        for row_index, row in enumerate(rows):
            if row_index == 0:  # 标题行
                title = ['使用人', '物品编码', '物品名称', '数量', '参考单价', '需求地点', '期望领用日期', '标准领用日期', '备注',
                         '配送方式', '验收人', '收货信息']
                if row != title:
                    return ['文件格式错误']
                continue
            apply_user = row[0]
            good_code = row[1]
            good_name = row[2]
            num = row[3]
            # price = row[4]
            # position = row[5]
            if file_type == 'xlsx':
                row[6] = row[6].date()
                require_date = row[6]
            elif file_type == 'xls':
                require_date = row[6]
            # standard_require_date = row[7]
            # remark = row[8]
            # delivery_method = row[9]
            # acceptor = row[10]
            # receiveInfo = row[11]

            pre_apply = {
                'apply_user': apply_user,
                'good_name': good_name,
                'good_code': good_code,
                'num': num,
                'require_date': require_date
            }
            pre_apply_serializer = PreApplySerializers(data=pre_apply)
            if not pre_apply_serializer.is_valid():
                err_msg = get_error_message(pre_apply_serializer)
                row.append(err_msg)
                CANNOT_APPLY.append(row)
                continue
            validated_list.append(row)
        return validated_list

    body = request.body
    json_body = json.loads(body)
    username = request.user.username
    org_id = json_body.get('org_id', None)

    if not org_id:
        raise BusinessException(StatusEnums.ORG_INFO_ERROR)

    valid_user_in_the_org(org_id, username)

    # 获取组名拼接dir_path
    org_name = Organization.objects.filter(id=org_id).first().org_name

    dir_path = os.path.join(org_name, 'analysis_apply_excel')
    file, file_path = tool_get_import_file(body, dir_path, 'file', 'fileName')

    storage = BKRepoStorage()

    # 将file以base64格式译码
    with open(file_path, 'wb') as f:
        f.write(base64.b64decode(file))

    # 保存到云端
    with open(file_path, 'rb') as fp:
        storage.save(file_path, fp)

    # 获取文件类型，支持xlsx于xls
    file_type = file_path.split('/')[-1].split('.')[-1]

    # 存放问题数据
    CANNOT_APPLY = []
    success_list = []
    if file_type == 'xlsx':
        xlsx = load_workbook(file_path)
        table = xlsx.worksheets[0]

        # 取得excel文件数据
        rows = tool_get_xlsx_excel_data(table)

        # 删除项目本地文件
        if os.path.exists(file_path):
            os.remove(file_path)

        if len(rows) > 1:
            success_list = handle_excel_data(rows, CANNOT_APPLY)  # 处理数据
        else:
            raise BusinessException(StatusEnums.IMPORT_FILE_EMPTY_ERROR)

    elif file_type == 'xls':
        xls = xlrd.open_workbook(file_path)
        table = xls.sheets()[0]
        rows = []

        # 取得excel文件数据
        for row_idx in range(table.nrows):
            row = []
            for col_idx in range(table.ncols):
                value = table.cell(row_idx, col_idx).value
                if table.cell(row_idx, col_idx).ctype == 3:
                    date = xldate_as_tuple(value, 0)
                    value = datetime.datetime(*date).date()
                row.append(value)
            rows.append(row)

        # 删除项目本地文件
        if os.path.exists(file_path):
            os.remove(file_path)

        if len(rows) > 1:
            success_list = handle_excel_data(rows, CANNOT_APPLY)  # 处理数据
        else:
            raise BusinessException(StatusEnums.IMPORT_FILE_EMPTY_ERROR)

    if success_list == ['文件格式错误']:
        result = {
            "code": 400,
            "result": False,
            "message": success_list[0],
            "data": {}
        }
        return get_result(result)

    elif not CANNOT_APPLY:
        result = {
            "code": 200,
            "result": True,
            "message": "导入成功",
            "data": {
                'success_list': success_list
            }
        }
        return get_result(result)
    else:
        can_not_apply_file_url = generate_can_not_apply_excel(CANNOT_APPLY, username, org_name)
        result = {
            "code": StatusEnums.IMPORT_ERROR.code,
            "result": False,
            "message": "部分/全部excel数据" + StatusEnums.IMPORT_ERROR.errmsg,
            "data": {
                'created_fail_list': CANNOT_APPLY,
                'file_url': can_not_apply_file_url,
                'success_list': success_list
            }
        }
        return get_result(result)
