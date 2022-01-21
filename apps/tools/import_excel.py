import base64
import json
import os.path

import xlrd
from openpyxl import load_workbook
from django.http import HttpResponse
from django.views.decorators.http import require_POST

from apps.good_apply.models import Position
from apps.good_purchase.models import UserInfo, Good, GroupApply
# from apps.good_purchase.serializers import GroupApplySerializers
from apps.tools.param_check import get_error_message
from apps.tools.response import get_result
from apps.utils.enums import StatusEnums
from apps.utils.exceptions import BusinessException
from django.conf import settings


@require_POST
def import_excel(request):
    def handle_excel_data(rows, CANNOT_ADD):  # 处理传入的列表数据，判断是否加入部门所需物资表
        group_apply_create_list = []  # 存放GroupApply对象
        for row_index, row in enumerate(rows):
            if row_index == 0:  # 标题行跳过
                continue

            gapply_item = row

            username = gapply_item[0]
            good_code = gapply_item[1]

            # 判断商品是否存在于商品表中，判断用户名是否存在于用户表中
            if not Good.objects.filter(good_code=good_code, status=1).exists():
                CANNOT_ADD.append(good_code)
                continue
            if not UserInfo.objects.filter(username=username).exists():
                CANNOT_ADD.append(good_code)
                continue
            phone = UserInfo.objects.filter(username=username).first().phone

            num = gapply_item[2]

            # 判断数量是否为整形或浮点型，并且是否大于等于0
            if (not isinstance(num, int) and not isinstance(num, float)) or not num >= 0:
                CANNOT_ADD.append(good_code)
                continue

            position = gapply_item[3]

            # 判断地区是否存在于地区表中
            if not Position.objects.filter(name=position).exists():
                CANNOT_ADD.append(good_code)
                continue

            get_date = gapply_item[4]
            remarks = gapply_item[5]
            good_name = gapply_item[6]
            group_apply_create_list.append(GroupApply(good_code=good_code, num=num, username=username, position=position,
                                      phone=phone, status=4, remarks=remarks))

        # 若无问题数据
        if not CANNOT_ADD:
            GroupApply.objects.bulk_create(group_apply_create_list)

    body = request.body

    # 判空
    body = json.loads(body)
    if not body['file']:
        raise BusinessException(StatusEnums.PARAMS_ERROR)

    file = body['file']

    # 判空
    if not body['fileName']:
        raise BusinessException(StatusEnums.PARAMS_ERROR)

    file_name = body['fileName']
    dir_path = os.path.join(settings.MEDIA_ROOT, 'import_excel')

    # 检查文件夹是否存在
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

    # 拼接文件路径
    file_path = os.path.join(dir_path, file_name)

    # 将file以base64格式译码
    with open(file_path, 'wb') as f:
        f.write(base64.b64decode(file))

    # 获取文件类型，支持xlsx于xls
    file_Type = file_name.split('.')[-1]

    # 存放问题数据
    CANNOT_ADD = []
    if file_Type == 'xlsx':
        xlsx = load_workbook(file_path)
        table = xlsx.worksheets[0]
        rows = []

        # 取得excel文件数据
        for i in range(1, table.max_row + 1):
            row = [table.cell(i, index).value for index in range(1, table.max_column + 1)]
            rows.append(row)
        if len(rows) > 1:
            handle_excel_data(rows, CANNOT_ADD)  # 处理数据
        else:
            raise BusinessException(StatusEnums.IMPORT_FILE_EMPTY_ERROR)

    elif file_Type == 'xls':
        xls = xlrd.open_workbook(file_path)
        table = xls.sheets()[0]
        rows = []

        # 取得excel文件数据
        for row_idx in range(table.nrows):
            rows.append(table.row_values(row_idx))
        if len(rows) > 1:
            handle_excel_data(rows, CANNOT_ADD)  # 处理数据
        else:
            raise BusinessException(StatusEnums.IMPORT_FILE_EMPTY_ERROR)

    if not CANNOT_ADD:
        result = {
            "code": 200,
            "result": True,
            "message": "导入成功",
            "data": {}
        }
        return get_result(result)
    else:
        result = {
            "code": StatusEnums.IMPORT_ERROR.code,
            "result": True,
            "message": "部分/全部excel数据" + StatusEnums.IMPORT_ERROR.errmsg,
            "data": {
                'created_fail_list': CANNOT_ADD
            }
        }
        return get_result(result)

